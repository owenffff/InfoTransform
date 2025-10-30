"""
FastAPI application for InfoTransform
"""

import os
import logging
from datetime import datetime
from pathlib import Path
import pandas as pd
from openpyxl.styles import Font, PatternFill
from io import BytesIO, StringIO

from contextlib import asynccontextmanager
from fastapi import FastAPI, HTTPException, status
from fastapi.responses import JSONResponse, StreamingResponse
from fastapi.exceptions import RequestValidationError
from fastapi.middleware.cors import CORSMiddleware
from starlette.requests import Request
import uvicorn

from infotransform.config import config
from infotransform.processors import (
    VisionProcessor,
    AudioProcessor,
    BatchProcessor,
    StructuredAnalyzerAgent,
)
from infotransform.api.document_transform_api import transform, shutdown_processor
from infotransform.api.review_api import router as review_router

# Setup logger
logger = logging.getLogger(__name__)

# Initialize processors
vision_processor = None
audio_processor = None
batch_processor = None
structured_analyzer = None


def init_processors():
    """Initialize processors with error handling"""
    global vision_processor, audio_processor, batch_processor, structured_analyzer
    try:
        config.validate()
        vision_processor = VisionProcessor()
        audio_processor = AudioProcessor()
        batch_processor = BatchProcessor(vision_processor, audio_processor)
        structured_analyzer = StructuredAnalyzerAgent()
        return True
    except Exception as e:
        print(f"Error initializing processors: {e}")
        return False


@asynccontextmanager
async def lifespan(app: FastAPI):
    """Manage application lifespan"""
    # Startup
    if init_processors():
        print("[OK] Processors initialized successfully")

        # Log Azure Document Intelligence status
        if config.DOCINTEL_ENDPOINT:
            print("[OK] Azure Document Intelligence enabled for image-based PDFs")
            print(f"     Endpoint: {config.DOCINTEL_ENDPOINT}")
        else:
            print("[INFO] Azure Document Intelligence not configured")
            print("       Image-based/scanned PDFs may fail to process")
            print("       See README.md for Azure setup instructions")
    else:
        print(
            "[ERROR] Failed to initialize processors. Please check your configuration."
        )

    yield

    # Shutdown
    await shutdown_processor()  # Shutdown the optimized processor
    print("[OK] Cleanup completed")


# Initialize FastAPI app
app = FastAPI(
    title=config.get("app.name", "Information Transformer"),
    description=config.get(
        "app.description", "Transform any file type into structured, actionable data"
    ),
    version=config.get("app.version", "2.0.0"),
    lifespan=lifespan,
)

# Add CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Configure appropriately for production
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Get paths relative to project root
project_root = Path(__file__).parent.parent.parent

# Create necessary directories
os.makedirs(config.UPLOAD_FOLDER, exist_ok=True)
os.makedirs(config.TEMP_EXTRACT_DIR, exist_ok=True)


@app.get("/")
async def index():
    """Redirect to Next.js frontend"""
    from fastapi.responses import RedirectResponse

    # Get frontend port from environment variables
    frontend_port = os.getenv("PORT", "3000")
    return RedirectResponse(url=f"http://localhost:{frontend_port}", status_code=307)


@app.get("/api/models")
async def list_analysis_models():
    """List available document schemas"""
    if not structured_analyzer:
        raise HTTPException(status_code=503, detail="Analyzer not initialized")

    return {
        "models": structured_analyzer.get_available_models(),
        "ai_models": structured_analyzer.get_available_ai_models(),
    }


# Add the new optimized streaming endpoint
app.post("/api/transform")(transform)

# Include review API router
app.include_router(review_router)


@app.post("/api/download-results")
async def download_results(request: Request):
    """Download transformation results as Excel or CSV"""
    data = await request.json()
    payload_results = data.get("results")
    format_type = data.get("format", "excel")
    fields = data.get("fields", None)

    if payload_results is None:
        raise HTTPException(status_code=400, detail="No results to download")

    try:
        rows = []
        summary = data.get("summary")

        # Accept both payload shapes
        if isinstance(payload_results, list):
            # list of row dicts already normalized
            rows = payload_results
        elif isinstance(payload_results, dict):
            results_list = payload_results.get("results", [])
            summary = payload_results.get("summary", summary)
            for r in results_list:
                if r.get("status") == "success" and r.get("structured_data"):
                    row = {"filename": r.get("filename")}
                    row.update(r.get("structured_data", {}))
                    rows.append(row)
        else:
            raise HTTPException(status_code=400, detail="Invalid results payload")

        if not rows:
            raise HTTPException(
                status_code=400, detail="No successful results to export"
            )

        # Build DataFrame
        df = pd.DataFrame(rows)

        # Ensure filename is first and respect fields order if provided
        if isinstance(fields, list) and fields:
            ordered = ["filename"] + [f for f in fields if f != "filename"]
            remaining = [c for c in df.columns if c not in ordered]
            df = df.reindex(
                columns=[col for col in ordered + remaining if col in df.columns]
            )
        else:
            if "filename" in df.columns:
                cols = ["filename"] + [c for c in df.columns if c != "filename"]
                df = df.reindex(columns=cols)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        if format_type == "csv":
            # Return CSV
            csv_buf = StringIO()
            df.to_csv(csv_buf, index=False)
            csv_buf.seek(0)
            return StreamingResponse(
                csv_buf,
                media_type="text/csv",
                headers={
                    "Content-Disposition": f"attachment; filename=transform_results_{timestamp}.csv"
                },
            )

        # Default to Excel
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Results", index=False)

            worksheet = writer.sheets["Results"]

            # Format header row
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(fill_type="solid", fgColor="E0E0E0")

            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        value_length = (
                            len(str(cell.value)) if cell.value is not None else 0
                        )
                    except Exception as e:
                        logger.warning(f"Could not determine length of cell value: {e}")
                        value_length = 0
                    if value_length > max_length:
                        max_length = value_length
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

            # Optional summary sheet
            if isinstance(summary, dict) and len(rows) > 1:
                summary_df = pd.DataFrame([summary])
                summary_df.to_excel(writer, sheet_name="Summary", index=False)

                summary_sheet = writer.sheets["Summary"]
                for cell in summary_sheet[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(fill_type="solid", fgColor="E0E0E0")

                for column in summary_sheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            value_length = (
                                len(str(cell.value)) if cell.value is not None else 0
                            )
                        except Exception:
                            value_length = 0
                        if value_length > max_length:
                            max_length = value_length
                    adjusted_width = min(max_length + 2, 50)
                    summary_sheet.column_dimensions[
                        column_letter
                    ].width = adjusted_width

        output.seek(0)
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=transform_results_{timestamp}.xlsx"
            },
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error generating download: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/health")
async def health_check():
    """Health check endpoint"""
    return {
        "status": "healthy",
        "processors_initialized": all(
            [
                vision_processor is not None,
                audio_processor is not None,
                structured_analyzer is not None,
            ]
        ),
        "server": "FastAPI",
        "version": config.get("app.version", "2.0.0"),
    }


# Custom exception handler for validation errors
@app.exception_handler(RequestValidationError)
async def validation_exception_handler(request: Request, exc: RequestValidationError):
    errors = []
    for error in exc.errors():
        errors.append({"loc": error["loc"], "msg": error["msg"], "type": error["type"]})

    print(f"Validation errors: {errors}")

    return JSONResponse(
        status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
        content={
            "detail": errors,
            "body": str(exc.body) if hasattr(exc, "body") else None,
        },
    )


# Custom exception handler for better error responses
@app.exception_handler(Exception)
async def general_exception_handler(request: Request, exc: Exception):
    return JSONResponse(status_code=500, content={"success": False, "error": str(exc)})


if __name__ == "__main__":
    # Run the FastAPI app with Uvicorn
    print(f"[START] Starting server on http://localhost:{config.PORT}")
    print(f"[DOCS] API documentation available at http://localhost:{config.PORT}/docs")

    uvicorn.run(
        app,
        host=config.get("api.host", "0.0.0.0"),
        port=config.PORT,
        reload=True,
        log_level="info",
    )
